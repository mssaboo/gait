import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
import sys


args = sys.argv[1]
print(args)
data = pd.read_excel (r''+args+'_Cleaned.xlsx')

sampling_time = 60/len(data.Time)
limit = (60/sampling_time) - 5


print("Total Number of Samples:")
print(len(data.Time))
print("Test Limit:")
print(limit)
print("Sampling Time:")
print(sampling_time)


#################################################################################################################################################

def Actual_HeelStrike (data,count=0):
	flag=0
	while(flag!=1):
		if((data.Heel[count]>=1)and(data.Heel[count-1]<1)and(data.Heel[count+1]>=1)and(data.Heel[count+2]>=1)and(data.Heel[count+3]>=1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def Actual_ToeStrike (data,count=0):
	flag=0
	limit_TS = count+(limit/100)
	while(flag!=1):
		if(data.Toe[count]>=1):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TS):
			flag=1
	return count

def Actual_HeelOff(data,count=0):
	flag=0
	limit_HO = count+(limit/50)
	while(flag!=1):
		if((data.Heel[count]==0)and(data.Heel[count-1]>0)and(data.Heel[count+1]==0)and(data.Heel[count+2]==0)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_HO):
			flag=1
	return count
	
def Actual_ToeOff(data,count=0):
	flag=0
	limit_TO = count+(limit/50)
	while(flag!=1):
		if((data.Toe[count]<1)and(data.Toe[count+1]<1)and(data.Toe[count+2]<1)and(data.Toe[count+3]<1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TO):
			flag=1
	return count
	
#################################################################################################################################################
A_HS=[]
A_TS=[]
A_HO=[]
A_TO=[]
count = 1

while(count<=limit):
	count = Actual_HeelStrike(data,count)
	A_HS.append(count)
	count=count+10
	
i=0
while(i<len(A_HS)):
	count = A_HS[i]+5
	if(count>=limit):
		break
	count = Actual_ToeStrike(data,count)
	A_TS.append(count)
	i=i+1

i=0
while(i<len(A_HS)):
	count = A_HS[i]
	count = Actual_HeelOff(data,count)
	A_HO.append(count)
	i=i+1

i=0
while(i<len(A_TS)):
	count = A_TS[i]
	count = Actual_ToeOff(data,count)
	A_TO.append(count)
	i=i+1


print("Actual Heel Strikes:")
print(A_HS)
print("Actual Toe Strikes:")
print(A_TS)
print("Actual Heel Offs:")
print(A_HO)
print("Actual Toe Offs:")
print(A_TO)
print("Number of Actual Heel Strikes:")
print(len(A_HS))
print("Number of Actual Toe Strikes:")
print(len(A_TS))
print("Number of Actual Heel Offs:")
print(len(A_HO))
print("Number of Actual Toe Offs:")
print(len(A_TO))

#############################################################################################################################################


def HS_fun1 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]>=50):
			if(data.SAVZ[count+1]<=data.SAVZ[count]):
				flag=1
			else:
				count=count+1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def HS_fun2 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.AANX[count]>=data.AANX[count+1]):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def HS_fun3(data,count=0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]<-30):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
		


def TS_fun1 (data,count=0):
	flag=0
	limit_TS=count+(limit/100)
	while(flag!=1):
		a=abs(data.AANY[count+1]-data.AANY[count])
		b=abs(data.AANY[count]-data.AANY[count-1])
		c=abs(data.AANY[count-1]-data.AANY[count-2])
		if((a<=0.55)and(b<=0.55)and(c<=0.55)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TS):
			flag=1
	return count


def HO_fun1 (data,count=0):
	flag=0
	limit_HO=count+(limit/150)
	while(flag!=1):
		a=abs(data.AANY[count+1]-data.AANY[count])
		if(a>=1):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_HO):
			flag=1
	return count


def TO_fun1 (data,count=0):
	flag=0
	limit_TO=count+(limit/150)
	while(flag!=1):
		a=data.AANY[count]
		if(a<-10.5):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TO):
			flag=1
	return count
#################################################################################################################################################

HS=[]
TS=[]
HO=[]
TO=[]
count =0

while(count<=limit):
	count = HS_fun1(data,count)
	count = count+1
	count = HS_fun2(data,count)
	count = count+1
	count = HS_fun3(data,count)
	HS.append(count)
	if(count>=limit):
		break
	count = count+1
	count = TS_fun1(data,count)
	
	TS.append(count)
	if(count>=limit):
		break	
	count = HO_fun1(data,count)
	HO.append(count)
	if(count>=limit):
		break	
	count = TO_fun1(data,count)
	TO.append(count)

print("Detected Heel Strikes:")
print(HS)
print("Detected Toe Strikes:")
print(TS)
print("Detected Heel Offs:")
print(HO)
print("Detected Toe Offs:")
print(TO)
print("Number of Detected Heel Strikes:")
print(len(HS))
print("Number of Detected Toe Strikes:")
print(len(TS))
print("Number of Detected Heel Offs:")
print(len(HO))
print("Number of Detected Toe Offs:")
print(len(TO))

#################################################################################################################################################

freq = 0
if (len(HS)>=len(A_HS)):
	freq = len(A_HS)
else:
	freq = len(HS)
diff_HS = []
count  = 0
total = 0
while(count<freq):
	diff_HS.append(abs(HS[count] - A_HS[count]))
	total = total + diff_HS[count]
	count = count+1
	
print("Difference in Heel Strike Actual and Detected:")
print(diff_HS)
print("Avg Diff in HS (Samples):")
print(total/freq)
print("Avg Diff in HS (Seconds):")
print(total*sampling_time/freq)


#################################################################################################################################################

freq = 0
if (len(TS)>=len(A_TS)):
	freq = len(A_TS)
else:
	freq = len(TS)
diff_TS = []
count  = 0
total = 0
while(count<freq):
	diff_TS.append(abs(TS[count] - A_TS[count]))
	total = total + diff_TS[count]
	count = count+1
	
print("Difference in Toe Strike Actual and Detected:")
print(diff_TS)
print("Avg Diff in TS (Samples):")
print(total/freq)
print("Avg Diff in TS (Seconds):")
print(total*sampling_time/freq)

#################################################################################################################################################
freq = 0
if (len(HO)>=len(A_HO)):
	freq = len(A_HO)
else:
	freq = len(HO)
diff_HO = []
count  = 0
total = 0
while(count<freq):
	diff_HO.append(abs(HO[count] - A_HO[count]))
	total = total + diff_HO[count]
	count = count+1
	
print("Difference in Heel Offs Actual and Detected:")
print(diff_HO)
print("Avg Diff in HO (Samples):")
print(total/freq)
print("Avg Diff in HO (Seconds):")
print(total*sampling_time/freq)

#################################################################################################################################################

freq = 0
if (len(TO)>=len(A_TO)):
	freq = len(A_TO)
else:
	freq = len(TO)
diff_TO = []
count  = 0
total = 0
while(count<freq):
	diff_TO.append(abs(TO[count] - A_TO[count]))
	total = total + diff_TO[count]
	count = count+1
	
print("Difference in Toe Offs Actual and Detected:")
print(diff_TO)
print("Avg Diff in TO (Samples):")
print(total/freq)
print("Avg Diff in TO (Seconds):")
print(total*sampling_time/freq)

#################################################################################################################################################

writer = pd.ExcelWriter(''+args+'_Tested.xlsx', engine='xlsxwriter')
data.to_excel(writer, sheet_name='Sheet1')
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
writer.save()

file = ''+args+'_Tested.xlsx'
wb = load_workbook(filename=file)
ws = wb.get_sheet_by_name('Sheet1')

for x in range(len(HS)):
	HS[x] = int(HS[x]) + 2
for x in range(len(TS)):
	TS[x] = int(TS[x]) + 2
for x in range(len(HO)):
	HO[x] = int(HO[x]) + 2
for x in range(len(TO)):
	TO[x] = int(TO[x]) + 2

if 'red_italic' not in wb.named_styles:
    red_italic = NamedStyle(name="red_italic")
    red_italic.font = Font(color='00FF0000', italic=True)
    wb.add_named_style(red_italic)

if 'green_italic' not in wb.named_styles:
    green_italic = NamedStyle(name="green_italic")
    green_italic.font = Font(color='00DF60', italic=True)
    wb.add_named_style(green_italic)


if 'dark_green_italic' not in wb.named_styles:
    dark_green_italic = NamedStyle(name="dark_green_italic")
    dark_green_italic.font = Font(color='fff68f', italic=True)
    wb.add_named_style(dark_green_italic)

if 'light_green_italic' not in wb.named_styles:
    light_green_italic = NamedStyle(name="light_green_italic")
    light_green_italic.font = Font(color='403a52', italic=True)
    wb.add_named_style(light_green_italic)


for x in HS:
	for cell in ws[x:x]:
    		cell.style = 'red_italic'

for x in TS:
	for cell in ws[x:x]:
    		cell.style = 'green_italic'

for x in HO:
	for cell in ws[x:x]:
    		cell.style = 'dark_green_italic'

for x in TO:
	for cell in ws[x:x]:
    		cell.style = 'light_green_italic'

wb.save(filename=file)

#################################################################################################################################################

