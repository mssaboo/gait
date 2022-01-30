import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
import sys
import csv


args = sys.argv[1]
print(args)
data = pd.read_excel (r''+args+'_Cleaned.xlsx')

sampling_time = 60/len(data.Time)
limit = (60/sampling_time) - 5


print("Total Number of Samples:")
print(len(data.Time))
print("Test Limit:")
print(limit)
print("Sampling Time:")
print(sampling_time)

#################################################################################################################################################

def Actual_HeelStrike (data,count=0):
	flag=0
	while(flag!=1):
		if((data.Heel[count]>=1)and(data.Heel[count-1]<1)and(data.Heel[count+1]>=1)and(data.Heel[count+2]>=1)and(data.Heel[count+3]>=1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def Actual_HeelMax (data,count=0):
	flag=0
	limit_HM = count+40
	maxi=0
	maxi_idx=count
	while(count<limit_HM):
		if(data.Heel[count]>maxi):
			maxi=data.Heel[count]
			maxi_idx = count
			count = count+1
		else:
			count=count+1
		if(count>=limit):
			break
	return maxi_idx

def Actual_HeelOff(data,count=0):
	flag=0
	while(flag!=1):
		if((data.Heel[count]==0)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
	
def Actual_ToeOff(data,count=0):
	flag=0
	while(flag!=1):
		if((data.Toe[count]<1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
	
#################################################################################################################################################
A_HS=[]
A_HM=[]
A_HO=[]
A_TO=[]
count = 0

while(count<=limit):
	count = Actual_HeelStrike(data,count)
	A_HS.append(count)
	count=count+10
	

i=0
while(i<len(A_HS)):
	count = A_HS[i]
	count = Actual_HeelOff(data,count)
	A_HO.append(count)
	i=i+1

i=0
while(i<len(A_HO)):
	count = A_HO[i]+5
	if(count>=limit):
		break
	count = Actual_ToeOff(data,count)
	A_TO.append(count)
	i=i+1

i=0
while(i<len(A_HO)):
	count = A_HS[i]
	count = Actual_HeelMax(data,count)
	A_HM.append(count)	
	i=i+1


#############################################################################################################################################


#################################################################################################################################################

print("Actual HS:")
print(len(A_HS))
print(A_HS)

print("Actual HM:")
print(len(A_HM))

print("Actual HO:")
print(len(A_HO))

print("Actual TO:")
print(len(A_TO))
#################################################################################################################################################

writer = pd.ExcelWriter(''+args+'_Actual.xlsx', engine='xlsxwriter')
data.to_excel(writer, sheet_name='Sheet1')
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
writer.save()

file = ''+args+'_Actual.xlsx'
wb = load_workbook(filename=file)
ws = wb.get_sheet_by_name('Sheet1')

for x in range(len(A_HS)):
	A_HS[x] = int(A_HS[x]) + 2
for x in range(len(A_HM)):
	A_HM[x] = int(A_HM[x]) + 2
for x in range(len(A_HO)):
	A_HO[x] = int(A_HO[x]) + 2
for x in range(len(A_TO)):
	A_TO[x] = int(A_TO[x]) + 2

if 'red_italic' not in wb.named_styles:
    red_italic = NamedStyle(name="red_italic")
    red_italic.font = Font(color='00FF0000', italic=True)
    wb.add_named_style(red_italic)

if 'green_italic' not in wb.named_styles:
    green_italic = NamedStyle(name="green_italic")
    green_italic.font = Font(color='00DF60', italic=True)
    wb.add_named_style(green_italic)


if 'dark_green_italic' not in wb.named_styles:
    dark_green_italic = NamedStyle(name="dark_green_italic")
    dark_green_italic.font = Font(color='fff68f', italic=True)
    wb.add_named_style(dark_green_italic)

if 'light_green_italic' not in wb.named_styles:
    light_green_italic = NamedStyle(name="light_green_italic")
    light_green_italic.font = Font(color='403a52', italic=True)
    wb.add_named_style(light_green_italic)


for x in A_HS:
	for cell in ws[x:x]:
    		cell.style = 'red_italic'

for x in A_HM:
	for cell in ws[x:x]:
    		cell.style = 'green_italic'

for x in A_HO:
	for cell in ws[x:x]:
    		cell.style = 'dark_green_italic'

for x in A_TO:
	for cell in ws[x:x]:
    		cell.style = 'light_green_italic'

wb.save(filename=file)
wb.close()
