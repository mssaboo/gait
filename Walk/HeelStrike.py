import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
import sys


args = sys.argv[1]
print(args)
data = pd.read_excel (r''+args+'_Cleaned.xlsx')
print(len(data.Time))
sampling_time = 60/len(data.Time)
limit = (60/sampling_time) - 5
print(limit)
print(sampling_time)

def fun1 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]>=50):
			if(data.SAVZ[count+1]<=data.SAVZ[count]):
				flag=1
			else:
				count=count+1
		else:
			count=count+1
			#print(count)
		if(count>=limit):
			break
	#print("hey")
	return count


def fun2 (data,count=0):
	flag=0
	while(flag!=1):
		if(data.AAVY[count]>0):
			if(data.AAVY[count+1]<=data.AAVY[count]):
				if(data.AAVY[count-1]<=data.AAVY[count]):
					flag=1
				else:
					count=count+1

			else:
				count=count+1

		elif(data.AAVX[count]>0):
			if(data.AAVX[count+1]<=data.AAVX[count]):
				if(data.AAVX[count-1]<=data.AAVX[count]):
					flag=1
				else:
					count=count+1

			else:
				count=count+1

		else:
			count = count+1
		if(count>=limit):
			break
	return count

def fun3(data,count=0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]<-30):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
		

def fun4 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.AANX[count]>=data.AANX[count+1]):
			flag=1
		else:
			count=count+1
			#print(count)
		if(count>=limit):
			break
	#print("hey")
	return count
ans=[]
count = 0

while(count<=limit):
#	print(data.SAVZ[1])
	count = fun1(data,count)
	count=count+1
	count = fun4(data,count)
	count=count+1
	count = fun3(data,count)
#	count = accy(data,count)
	ans.append(count)
	count = count+10
print(ans)

def HeelStrike (data, count =0):
	flag=0
	while(flag!=1):
		if((data.Heel[count]>=1)and(data.Heel[count-1]<1)and(data.Heel[count+1]>=1)and(data.Heel[count+2]>=1)and(data.Heel[count+3]>=1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


ans_original=[]
count = 1

while(count<=limit):
	count = HeelStrike(data,count)
	ans_original.append(count)
	count=count+1
print(ans_original)
freq = 0
if (len(ans)>=len(ans_original)):
	freq = len(ans_original)
else:
	freq = len(ans)
diff = []
count  = 0
total = 0
while(count<freq):
	diff.append(abs(ans[count] - ans_original[count]))
	total = total + diff[count]
	count = count+1
	
print(diff)
print(len(ans))
print(len(ans_original))
print(len(diff))
print(total/freq)
print(total*sampling_time/freq)
writer = pd.ExcelWriter(''+args+'_Tested.xlsx', engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object.
data.to_excel(writer, sheet_name='Sheet1')

# Get the xlsxwriter objects from the dataframe writer object.
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
writer.save()

file = ''+args+'_Tested.xlsx'
wb = load_workbook(filename=file)
#ws = wb['Sheet1']
#red_font = Font(color='00FF0000', italic=True)

# Enumerate the cells in the second row
#for i in range(len(ans)):
for x in range(len(ans)):
	ans[x] = int(ans[x]) + 2
#print(ans)
#for i in ans:
#	for cell in ws["i:i"]:
#		cell.font = red_font

#wb.save(filename=file)


ws = wb.get_sheet_by_name('Sheet1')



# Create a NamedStyle (if not already defined)
if 'red_italic' not in wb.named_styles:
    red_italic = NamedStyle(name="red_italic")
    red_italic.font = Font(color='00FF0000', italic=True)
    wb.add_named_style(red_italic)

# Enumerate the cells in the second row
for x in ans:
	for cell in ws[x:x]:
    		cell.style = 'red_italic'

wb.save(filename=file)

