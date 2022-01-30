import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
import sys
import csv


args = sys.argv[1]
print(args)
data = pd.read_excel (r''+args+'_Cleaned.xlsx')

sampling_time = 60/len(data.Time)
limit = (60/sampling_time) - 5


print("Total Number of Samples:")
print(len(data.Time))
print("Test Limit:")
print(limit)
print("Sampling Time:")
print(sampling_time)


#################################################################################################################################################

def Actual_HeelStrike (data,count=0):
	flag=0
	while(flag!=1):
		if((data.Heel[count]>=1)and(data.Heel[count-1]<1)and(data.Heel[count+1]>=1)and(data.Heel[count+2]>=1)and(data.Heel[count+3]>=1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def Actual_HeelMax (data,count=0):
	flag=0
	while(flag!=1):
		if(data.Heel[count]>data.Heel[count+1]):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def Actual_HeelOff(data,count=0):
	flag=0
	limit_HO = count+(limit/50)
	while(flag!=1):
		if((data.Heel[count]==0)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_HO):
			flag=1
	return count
	
def Actual_ToeOff(data,count=0):
	flag=0
	limit_TO = count+(limit/50)
	while(flag!=1):
		if((data.Toe[count]<1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TO):
			flag=1
	return count
	
#################################################################################################################################################
A_HS=[]
A_HM=[]
A_HO=[]
A_TO=[]
count = 1

while(count<=limit):
	count = Actual_HeelStrike(data,count)
	A_HS.append(count)
	count=count+10
	
i=0
while(i<len(A_HS)):
	count = A_HS[i]+7
	if(count>=limit):
		break
	count = Actual_HeelMax(data,count)
	A_HM.append(count)
	i=i+1

i=0
while(i<len(A_HS)):
	count = A_HS[i]
	count = Actual_HeelOff(data,count)
	A_HO.append(count)
	i=i+1

i=0
while(i<len(A_HO)):
	count = A_HO[i]+5
	if(count>=limit):
		break
	count = Actual_ToeOff(data,count)
	A_TO.append(count)
	i=i+1



#############################################################################################################################################


def HS_fun1 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]>=50):
			if(data.SAVZ[count+1]<=data.SAVZ[count]):
				flag=1
			else:
				count=count+1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def HS_fun2 (data, count =0):
	flag=0
	while(flag!=1):
		if(data.AANX[count]>=data.AANX[count+1]):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def HS_fun3(data,count=0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]<-30):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
		


def HM_fun1 (data,count=0):
	flag=0
	limit_HM = count+(limit/150)
	while(flag!=1):
		a=abs(data.AANY[count+1]-data.AANY[count])
		if((a==0)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_HM):
			flag=1
	return count


def HO_fun1 (data,count=0):
	flag=0
	limit_HO=count+(limit/100)
	while(flag!=1):
		#a=abs(data.AANY[count+1]-data.AANY[count])
		#if(data.SANY[count]>data.SANY[count+1]):
		#	flag=1
		if(((data.AAVY[count+2]-data.AAVY[count+1])>=1)and((data.AAVY[count+1]-data.AAVY[count])>=1)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_HO):
			flag=1
	return count


def TO_fun1 (data,count=0):
	flag=0
	limit_TO=count+(limit/150)
	while(flag!=1):
		a=data.AANY[count]
		b=data.SAVZ[count]
		if((a<-10.5)and(b<-111)):
			flag=1
		else:
			count=count+1
		if(count>=limit):
			break
		if(count>=limit_TO):
			flag=1
	return count
#################################################################################################################################################

HS=[]
HM=[]
HO=[]
TO=[]
count =0

while(count<=limit):
	count = HS_fun1(data,count)
	count = count+1
	count = HS_fun2(data,count)
	count = count+1
	count = HS_fun3(data,count)
	HS.append(count)
	if(count>=limit):
		break
	count = count+5
	count = HM_fun1(data,count)
	
	HM.append(count)
	count = count+5
	if(count>=limit):
		break	
	count = HO_fun1(data,count)
	HO.append(count)
	if(count>=limit):
		break	
	count = TO_fun1(data,count)
	TO.append(count)



#################################################################################################################################################

freq = 0
if (len(HS)>=len(A_HS)):
	freq = len(A_HS)
else:
	freq = len(HS)
diff_HS = []
count  = 0
total = 0
while(count<freq):
	diff_HS.append(abs(HS[count] - A_HS[count]))
	total = total + diff_HS[count]
	count = count+1
print("Actual Heel Strikes:")
print(A_HS)
print("Detected Heel Strikes:")
print(HS)
print("Number of Actual Heel Strikes:")
print(len(A_HS))
print("Number of Detected Heel Strikes:")
print(len(HS))	
print("Difference in Heel Strike Actual and Detected:")
print(diff_HS)
print("Avg Diff in HS (Samples):")
print(total/freq)
print("Avg Diff in HS (Seconds):")
print(total*sampling_time/freq)
HS_error = total*sampling_time/freq

#################################################################################################################################################

freq = 0
if (len(HM)>=len(A_HM)):
	freq = len(A_HM)
else:
	freq = len(HM)
diff_HM = []
count  = 0
total = 0
while(count<freq):
	diff_HM.append(abs(HM[count] - A_HM[count]))
	total = total + diff_HM[count]
	count = count+1

print("Actual HeelMax:")
print(A_HM)
print("Detected HeelMax:")
print(HM)	
print("Number of Actual HeelMax:")
print(len(A_HM))
print("Number of Detected HeelMax:")
print(len(HM))
print("Difference in HeelMax Actual and Detected:")
print(diff_HM)
print("Avg Diff in HM (Samples):")
print(total/freq)
print("Avg Diff in HM (Seconds):")
print(total*sampling_time/freq)
HM_error = total*sampling_time/freq

#################################################################################################################################################
freq = 0
if (len(HO)>=len(A_HO)):
	freq = len(A_HO)
else:
	freq = len(HO)
diff_HO = []
count  = 0
total = 0
while(count<freq):
	diff_HO.append(abs(HO[count] - A_HO[count]))
	total = total + diff_HO[count]
	count = count+1

print("Actual Heel Offs:")
print(A_HO)
print("Detected Heel Offs:")
print(HO)	
print("Number of Actual Heel Offs:")
print(len(A_HO))
print("Number of Detected Heel Offs:")
print(len(HO))
print("Difference in Heel Offs Actual and Detected:")
print(diff_HO)
print("Avg Diff in HO (Samples):")
print(total/freq)
print("Avg Diff in HO (Seconds):")
print(total*sampling_time/freq)
HO_error = total*sampling_time/freq
#################################################################################################################################################

freq = 0
if (len(TO)>=len(A_TO)):
	freq = len(A_TO)
else:
	freq = len(TO)
diff_TO = []
count  = 0
total = 0
while(count<freq):
	diff_TO.append(abs(TO[count] - A_TO[count]))
	total = total + diff_TO[count]
	count = count+1

print("Actual Toe Offs:")
print(A_TO)
print("Detected Toe Offs:")
print(TO)
print("Number of Actual Toe Offs:")
print(len(A_TO))
print("Number of Detected Toe Offs:")
print(len(TO))	
print("Difference in Toe Offs Actual and Detected:")
print(diff_TO)
print("Avg Diff in TO (Samples):")
print(total/freq)
print("Avg Diff in TO (Seconds):")
print(total*sampling_time/freq)
TO_error = total*sampling_time/freq

#################################################################################################################################################

print("Avg Diff in HS (Seconds):")
print(HS_error)
print("Avg Diff in HM (Seconds):")
print(HM_error)
print("Avg Diff in HO (Seconds):")
print(HO_error)
print("Avg Diff in TO (Seconds):")
print(TO_error)

#################################################################################################################################################


writer = pd.ExcelWriter(''+args+'_Tested.xlsx', engine='xlsxwriter')
data.to_excel(writer, sheet_name='Sheet1')
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
writer.save()

file = ''+args+'_Tested.xlsx'
wb = load_workbook(filename=file)
ws = wb.get_sheet_by_name('Sheet1')

for x in range(len(HS)):
	HS[x] = int(HS[x]) + 2
for x in range(len(HM)):
	HM[x] = int(HM[x]) + 2
for x in range(len(HO)):
	HO[x] = int(HO[x]) + 2
for x in range(len(TO)):
	TO[x] = int(TO[x]) + 2

if 'red_italic' not in wb.named_styles:
    red_italic = NamedStyle(name="red_italic")
    red_italic.font = Font(color='00FF0000', italic=True)
    wb.add_named_style(red_italic)

if 'green_italic' not in wb.named_styles:
    green_italic = NamedStyle(name="green_italic")
    green_italic.font = Font(color='00DF60', italic=True)
    wb.add_named_style(green_italic)


if 'dark_green_italic' not in wb.named_styles:
    dark_green_italic = NamedStyle(name="dark_green_italic")
    dark_green_italic.font = Font(color='fff68f', italic=True)
    wb.add_named_style(dark_green_italic)

if 'light_green_italic' not in wb.named_styles:
    light_green_italic = NamedStyle(name="light_green_italic")
    light_green_italic.font = Font(color='403a52', italic=True)
    wb.add_named_style(light_green_italic)


for x in HS:
	for cell in ws[x:x]:
    		cell.style = 'red_italic'

for x in HM:
	for cell in ws[x:x]:
    		cell.style = 'green_italic'

for x in HO:
	for cell in ws[x:x]:
    		cell.style = 'dark_green_italic'

for x in TO:
	for cell in ws[x:x]:
    		cell.style = 'light_green_italic'

wb.save(filename=file)

#################################################################################################################################################

fc = open('Errors.txt','a')
fc.write(args + ' ' + str(HS_error) +' '+  str(HM_error) +' '+ str(HO_error) +' ' + str(TO_error) +'\n')
fc.close()

###############################################################################################################################################

