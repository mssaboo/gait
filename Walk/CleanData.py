import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
import sys


args = sys.argv[1]
print(args)
data = pd.read_excel (r''+args+'.xlsx')
print(len(data.Time))
sampling_time = 60/len(data.Time)
limit = (60/sampling_time) - 5
print(limit)
list_av = ['SAVZ','SAVY','SAVX','AAVZ','AAVY','AAVX']
list_ac = ['SACZ','SACY','SACX','AACZ','AACY','AACX']
list_an = ['SANZ','SANY','SANX','AANZ','AANY','AANX']

def SAVZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SAVZ[count]>=500):
			flag = 1
		elif(data.SAVZ[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SAVY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SAVY[count]>=500):
			flag = 1
		elif(data.SAVY[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SAVX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SAVX[count]>=500):
			flag = 1
		elif(data.SAVX[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def AAVZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AAVZ[count]>=500):
			flag = 1
		elif(data.AAVZ[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count



def AAVY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AAVY[count]>=500):
			flag = 1
		elif(data.AAVY[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def AAVX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AAVX[count]>=500):
			flag = 1
		elif(data.AAVX[count]<=-500):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SANZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SANZ[count]>=190):
			flag = 1
		elif(data.SANZ[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SANY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SANY[count]>=190):
			flag = 1
		elif(data.SANY[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SANX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SANX[count]>=190):
			flag = 1
		elif(data.SANX[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def AANZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AANZ[count]>=190):
			flag = 1
		elif(data.AANZ[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count



def AANY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AANY[count]>=190):
			flag = 1
		elif(data.AANY[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def AANX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AANX[count]>=190):
			flag = 1
		elif(data.AANX[count]<=-190):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count
def SACZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SACZ[count]>=200):
			flag = 1
		elif(data.SACZ[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SACY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SACY[count]>=200):
			flag = 1
		elif(data.SACY[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def SACX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.SACX[count]>=200):
			flag = 1
		elif(data.SACX[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count

def AACZ (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AACZ[count]>=200):
			flag = 1
		elif(data.AACZ[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count



def AACY (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AACY[count]>=200):
			flag = 1
		elif(data.AACY[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


def AACX (data, count = 0):
	flag=0
	while(flag!=1):
		if(data.AACX[count]>=200):
			flag = 1
		elif(data.AACX[count]<=-200):
			flag = 1
		else:
			count=count+1
		if(count>=limit):
			break
	return count


ans_SAVZ=[]
ans_SAVY=[]
ans_SAVX=[]
ans_SANZ=[]
ans_SANY=[]
ans_SANX=[]
ans_SACZ=[]
ans_SACY=[]
ans_SACX=[]
ans_AAVZ=[]
ans_AAVY=[]
ans_AAVX=[]
ans_AANZ=[]
ans_AANY=[]
ans_AANX=[]
ans_AACZ=[]
ans_AACY=[]
ans_AACX=[]

count = 0

while(count<=limit):
	count = SAVZ(data,count)
	ans_SAVZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = SAVY(data,count)
	ans_SAVY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = SAVX(data,count)
	ans_SAVX.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = SANZ(data,count)
	ans_SANZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = SANY(data,count)
	ans_SANY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = SANX(data,count)
	ans_SANX.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = SACZ(data,count)
	ans_SACZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = SACY(data,count)
	ans_SACY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = SACX(data,count)
	ans_SACX.append(count)
	count = count+1


count = 0
while(count<=limit):
	count = AAVZ(data,count)
	ans_AAVZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = AAVY(data,count)
	ans_AAVY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = AAVX(data,count)
	ans_AAVX.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = AANZ(data,count)
	ans_AANZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = AANY(data,count)
	ans_AANY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = AANX(data,count)
	ans_AANX.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = AACZ(data,count)
	ans_AACZ.append(count)
	count = count+1
count = 0

while(count<=limit):
	count = AACY(data,count)
	ans_AACY.append(count)
	count = count+1

count = 0

while(count<=limit):
	count = AACX(data,count)
	ans_AACX.append(count)
	count = count+1




writer = pd.ExcelWriter(''+args+'_Cleaned.xlsx', engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object.
data.to_excel(writer, sheet_name='Sheet1')

# Get the xlsxwriter objects from the dataframe writer object.
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
writer.save()

file = ''+args+'_Cleaned.xlsx'
wb = load_workbook(filename=file)
#ws = wb['Sheet1']
#red_font = Font(color='00FF0000', italic=True)

# Enumerate the cells in the second row
#for i in range(len(ans_SAVZ)):
for x in range(len(ans_SAVZ)):
	ans_SAVZ[x] = int(ans_SAVZ[x]) + 2
for x in range(len(ans_SAVY)):
	ans_SAVY[x] = int(ans_SAVY[x]) + 2
for x in range(len(ans_SAVX)):
	ans_SAVX[x] = int(ans_SAVX[x]) + 2
for x in range(len(ans_SANZ)):
	ans_SANZ[x] = int(ans_SANZ[x]) + 2
for x in range(len(ans_SANY)):
	ans_SANY[x] = int(ans_SANY[x]) + 2
for x in range(len(ans_SANX)):
	ans_SANX[x] = int(ans_SANX[x]) + 2
for x in range(len(ans_SACZ)):
	ans_SACZ[x] = int(ans_SACZ[x]) + 2
for x in range(len(ans_SACY)):
	ans_SACY[x] = int(ans_SACY[x]) + 2
for x in range(len(ans_SACX)):
	ans_SACX[x] = int(ans_SACX[x]) + 2
for x in range(len(ans_AAVZ)):
	ans_AAVZ[x] = int(ans_AAVZ[x]) + 2
for x in range(len(ans_AAVY)):
	ans_AAVY[x] = int(ans_AAVY[x]) + 2
for x in range(len(ans_AAVX)):
	ans_AAVX[x] = int(ans_AAVX[x]) + 2
for x in range(len(ans_AANZ)):
	ans_AANZ[x] = int(ans_AANZ[x]) + 2
for x in range(len(ans_AANY)):
	ans_AANY[x] = int(ans_AANY[x]) + 2
for x in range(len(ans_AANX)):
	ans_AANX[x] = int(ans_AANX[x]) + 2
for x in range(len(ans_AACZ)):
	ans_AACZ[x] = int(ans_AACZ[x]) + 2
for x in range(len(ans_AACY)):
	ans_AACY[x] = int(ans_AACY[x]) + 2
for x in range(len(ans_AACX)):
	ans_AACX[x] = int(ans_AACX[x]) + 2

#print(ans_SAVZ)
#for i in ans_SAVZ:
#	for cell in ws["i:i"]:
#		cell.font = red_font

#wb.save(filename=file)


ws = wb.get_sheet_by_name('Sheet1')

# Create a NamedStyle (if not already defined)
if 'red_italic' not in wb.named_styles:
    red_italic = NamedStyle(name="red_italic")
    red_italic.font = Font(color='00FF0000', italic=True)
    wb.add_named_style(red_italic)

# Enumerate the cells in the second row
for x in ans_SAVZ:
	ws.cell(x,5).value=(ws.cell(x+1,5).value + ws.cell(x-1,5).value )/2
for x in ans_SAVY:
	ws.cell(x,6).value=(ws.cell(x+1,6).value + ws.cell(x-1,6).value )/2
for x in ans_SAVX:
	ws.cell(x,7).value=(ws.cell(x+1,7).value + ws.cell(x-1,7).value )/2
for x in ans_SANZ:
	ws.cell(x,8).value=(ws.cell(x+1,8).value + ws.cell(x-1,8).value )/2
for x in ans_SANY:
	ws.cell(x,9).value=(ws.cell(x+1,9).value + ws.cell(x-1,9).value )/2
for x in ans_SANX:
	ws.cell(x,10).value=(ws.cell(x+1,10).value + ws.cell(x-1,10).value )/2
for x in ans_SACZ:
	ws.cell(x,11).value=(ws.cell(x+1,11).value + ws.cell(x-1,11).value )/2
for x in ans_SACY:
	ws.cell(x,12).value=(ws.cell(x+1,12).value + ws.cell(x-1,12).value )/2
for x in ans_SACX:
	ws.cell(x,13).value=(ws.cell(x+1,13).value + ws.cell(x-1,13).value )/2
for x in ans_AAVZ:
	ws.cell(x,14).value=(ws.cell(x+1,14).value + ws.cell(x-1,14).value )/2
for x in ans_AAVY:
	ws.cell(x,15).value=(ws.cell(x+1,15).value + ws.cell(x-1,15).value )/2
for x in ans_AAVX:
	ws.cell(x,16).value=(ws.cell(x+1,16).value + ws.cell(x-1,16).value )/2
for x in ans_AANZ:
	ws.cell(x,17).value=(ws.cell(x+1,17).value + ws.cell(x-1,17).value )/2
for x in ans_AANY:
	ws.cell(x,18).value=(ws.cell(x+1,18).value + ws.cell(x-1,18).value )/2
for x in ans_AANX:
	ws.cell(x,19).value=(ws.cell(x+1,19).value + ws.cell(x-1,19).value )/2
for x in ans_AACZ:
	ws.cell(x,20).value=(ws.cell(x+1,20).value + ws.cell(x-1,20).value )/2
for x in ans_AACY:
	ws.cell(x,21).value=(ws.cell(x+1,21).value + ws.cell(x-1,21).value )/2
for x in ans_AACX:
	ws.cell(x,22).value=(ws.cell(x+1,22).value + ws.cell(x-1,22).value )/2


wb.save(filename=file)

